import openpyxl

#################################
subjects=[]
chapters=[]
topics=[]
#################################

# Load the workbook
workbook = openpyxl.load_workbook("NEET-PLAN.xlsx")

# Get the sheet names
sheet_names = workbook.sheetnames

for i in range(len(sheet_names)):
    subjects.append({'id':i+1,'name': sheet_names[i]})

print(sheet_names)

# Select the sheet you want to read from (e.g., the first sheet)
# sheet = workbook.active


def genNum():
    n=1
    while True:
        yield n
        n+=1

gen_num=genNum()

data=[]
chapter_id=0
for sub_id in range(len(sheet_names)):
    count=0
    sheet = workbook[sheet_names[sub_id]]
    for row in sheet.iter_rows(values_only=True):
        if count==0:
            chapter=[{"id":gen_num.__next__(),"chapter":row[i],"subject":sub_id+1} for i in range(len(row))]
            chapter_id+=len(row)
            count+=1
            continue
        chapters.extend(chapter)
        topics.extend([{"topic":row[i],"chap_id":chapter[i]["id"]} for i in range(len(row)) if row[i]!=None])
        data.append(row)


print(subjects)
print("\n\nchapter..................................",chapters)
print("\n\ntopic......................................",topics)


# sheet = workbook[sheet_names[0]]
# # Read the data from the sheet and store it in a list of lists
# data = []
# count=0
# for row in sheet.iter_rows(values_only=True):
#     if count==0:
#         chapter.extend([{"id":i+1,"chapter":row[i],"subject":1} for i in range(len(row))])
#         count+=1
#         continue
#     count+=1
#     topic.extend([{"id":i+1,"topic":row[i],"chap_id":chapter[i]["id"]} for i in range(len(row)) if row[i]!=None])
#     data.append(row)

# Now 'data' contains the content of the Excel sheet as a list of lists
# print(data)
